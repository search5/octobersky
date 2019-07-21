import datetime
import io

import paginate
from flask import request, url_for, render_template, flash, send_file
from flask.views import MethodView
from flask_login import login_required
from openpyxl import load_workbook
from paginate_sqlalchemy import SqlalchemyOrmWrapper
from sqlalchemy import desc

from nanumlectures.common import is_admin_role, paginate_link_tag
from nanumlectures.database import db_session
from nanumlectures.models import Roundtable, VoteBooks


class VoteBooksListView(MethodView):
    decorators = [is_admin_role, login_required]

    def get(self):
        current_page = request.args.get("page", 1, type=int)
        search_option = request.args.get("search_option", '')
        search_word = request.args.get("search_word", '')

        if search_option == "roundtable_num" and search_word and not search_word.isdecimal():
            flash('개최회차는 숫자만 입력하셔야 합니다.')
            search_word = None

        page_url = url_for("admin.vote_books")
        if search_word:
            page_url = url_for("admin.vote_books", search_option=search_option, search_word=search_word)
            page_url = str(page_url) + "&page=$page"
        else:
            page_url = str(page_url) + "?page=$page"

        items_per_page = 10

        records = db_session.query(VoteBooks).join(Roundtable)
        if search_word:
            if search_option == 'roundtable_num':
                records = records.filter(Roundtable.roundtable_num == search_word)
        records = records.order_by(desc(VoteBooks.id))
        total_cnt = records.count()

        paginator = paginate.Page(records, current_page, page_url=page_url,
                                  items_per_page=items_per_page,
                                  wrapper_class=SqlalchemyOrmWrapper)

        return render_template("admin/vote_books.html", paginator=paginator,
                               paginate_link_tag=paginate_link_tag,
                               page_url=page_url, items_per_page=items_per_page,
                               total_cnt=total_cnt, page=current_page)


class VoteBooksListCsvView(MethodView):
    decorators = [is_admin_role, login_required]

    def get(self):
        records = db_session.query(VoteBooks).join(Roundtable)
        records = records.order_by(desc(VoteBooks.id))

        wb = load_workbook(filename='nanumlectures/ext/octobersky_book_list.xlsx')
        ws = wb['시트1']

        def book_info_split(book_split):
            book_split.extend([None] * (3 - len(book_split)))
            return book_split

        for index, entry in enumerate(records, 4):
            ws["B{}".format(index)] = entry.lecture_user.name
            ws["C{}".format(index)] = entry.lecture.lecture_title

            book1 = book_info_split(entry.book_info['book1'].split(",", 3))
            ws["D{}".format(index)] = book1[0]
            ws["E{}".format(index)] = book1[1]
            ws["F{}".format(index)] = book1[2]
            ws["G{}".format(index)] = entry.book_info['book1_desc']

            book2 = book_info_split(entry.book_info['book2'].split(",", 3))
            ws["H{}".format(index)] = book2[0]
            ws["I{}".format(index)] = book2[1]
            ws["J{}".format(index)] = book2[2]

            book3 = book_info_split(entry.book_info['book3'].split(",", 3))
            ws["K{}".format(index)] = book3[0]
            ws["L{}".format(index)] = book3[1]
            ws["M{}".format(index)] = book3[2]

            book_etc = book_info_split(entry.book_info['etc'].split(",", 3))
            ws["N{}".format(index)] = book_etc[0]
            ws["O{}".format(index)] = book_etc[1]
            ws["P{}".format(index)] = book_etc[2]

        resultIO = io.BytesIO()
        wb.save(resultIO)
        resultIO.seek(0)

        file_name = '10월의 하늘_강연자 책 추천 리스트_{}.xlsx'.format(datetime.datetime.now().strftime("%Y-%m-%d"))

        return send_file(resultIO, mimetype="application/zip", as_attachment=True, attachment_filename=file_name)
