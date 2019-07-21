import datetime
import io

import paginate
from flask import render_template, request, url_for, send_file
from flask.views import MethodView
from flask_login import login_required
from openpyxl import Workbook
from paginate_sqlalchemy import SqlalchemyOrmWrapper
from sqlalchemy import desc

from nanumlectures.common import is_admin_role, paginate_link_tag
from nanumlectures.database import db_session
from nanumlectures.models import OTandParty
from nanumlectures.routes.public import library_find


class OTAndPartyView(MethodView):
    decorators = [is_admin_role, login_required]

    def get(self):
        current_page = request.args.get("page", 1, type=int)
        ot1_join = request.args.get("ot1_join", '')
        ot2_join = request.args.get("ot2_join", '')
        party_join = request.args.get("party_join", '')

        page_url = url_for("admin.otandparty")
        page_url = str(page_url) + "?page=$page"

        if ot1_join:
            page_url += '&ot1_join=' + ot1_join

        if ot2_join:
            page_url += '&ot2_join=' + ot2_join

        if party_join:
            page_url += '&party_join=' + party_join

        items_per_page = 10

        records = db_session.query(OTandParty)
        records = records.filter(OTandParty.ot1_join == (True if ot1_join == 'on' else False))
        records = records.filter(OTandParty.ot2_join == (True if ot2_join == 'on' else False))
        records = records.filter(OTandParty.party_join == (True if party_join == 'on' else False))

        records = records.order_by(desc(OTandParty.id))
        total_cnt = records.count()

        paginator = paginate.Page(records, current_page, page_url=page_url,
                                  items_per_page=items_per_page,
                                  wrapper_class=SqlalchemyOrmWrapper)

        return render_template("admin/ot_and_party.html", paginator=paginator,
                               paginate_link_tag=paginate_link_tag,
                               page_url=page_url, items_per_page=items_per_page,
                               total_cnt=total_cnt, page=current_page)


class OTAndPartyCsvView(MethodView):
    decorators = [is_admin_role, login_required]

    def get(self):
        records = db_session.query(OTandParty).order_by(desc(OTandParty.id))

        wb = Workbook()
        ws = wb.active

        ws['A1'] = '#'
        ws['B1'] = '도서관'
        ws['C1'] = '신청자'
        ws['D1'] = '1차 OT'
        ws['E1'] = '2차 OT'
        ws['F1'] = '뒤풀이'

        for index, entry in enumerate(records, 2):
            ws["A{}".format(index)] = index - 1
            ws["B{}".format(index)] = library_find(entry.party_user, entry.roundtable)
            ws["C{}".format(index)] = entry.party_user.name
            ws["D{}".format(index)] = entry.ot1_join and '참석' or '불참'
            ws["E{}".format(index)] = entry.ot2_join and '참석' or '불참'
            ws["F{}".format(index)] = entry.party_join and '참석' or '불참'

        resultIO = io.BytesIO()
        wb.save(resultIO)
        resultIO.seek(0)

        file_name = '10월의 하늘_OT및뒤풀이참석자_{}.xlsx'.format(datetime.datetime.now().strftime("%Y-%m-%d"))

        return send_file(resultIO, mimetype="application/zip", as_attachment=True, attachment_filename=file_name)
